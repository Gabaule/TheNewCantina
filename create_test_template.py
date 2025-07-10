import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_detailed_test_case_sheet():
    """
    Generates a single, detailed test case sheet in Excel format,
    replicating the provided image's structure and style.
    """
    # 1. --- Data for the Test Case (Populated from your project) ---
    test_case_data = {
        "id": "MANUAL-TC-001",
        "description": "Verify Visual Feedback and State on Balance Top-Up",
        "created_by": "Gabriel Aumasson-Leduc",
        "reviewed_by": "Ishan Baichoo",
        "version": "1.0",
        "qa_log": "Initial draft for manual testing phase. Covers HTMX out-of-band swaps.",
        "tester_name": "Clément De Simon",
        "date_tested": "2025-07-08",
        "final_status": "Pass",
        "prerequisites": [
            "User is logged in as 'student1@example.com'.",
            "Application is running in a modern web browser.",
        ],
        "test_data": [
            "Amount to add: 25.50",
        ],
        "scenario": "Verify that when a user adds funds, the UI provides clear feedback and all balance indicators update correctly without a full page reload.",
        "steps": [
            {"num": 1, "details": "Navigate to the 'Top Up Balance' page.", "expected": "The 'Account Balance' page is displayed.", "actual": "As Expected", "status": "Pass"},
            {"num": 2, "details": "Enter '25.50' into the amount field and click 'Add Money'.", "expected": "A success message appears and the 'Current Balance' on the page updates.", "actual": "As Expected", "status": "Pass"},
            {"num": 3, "details": "Observe the header of the application.", "expected": "The balance in the top-right corner updates instantly without a page refresh.", "actual": "As Expected", "status": "Pass"},
            {"num": 4, "details": "Navigate to the main 'Dashboard' page.", "expected": "The balance in the header remains at the new, updated value.", "actual": "As Expected", "status": "Pass"}
        ]
    }

    # 2. --- Define Styles ---
    header_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid") # Light yellow
    bold_font = Font(name='Calibri', bold=True)
    thin_border_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def apply_header_style(cell):
        cell.font = bold_font
        cell.fill = header_fill

    # 3. --- Create and Setup Workbook ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = test_case_data["id"]

    # 4. --- Build the Layout and Populate Data ---
    
    # Section 1: Top Header Info
    ws.merge_cells('A1:B1'); ws['A1'] = "Test Case ID"
    ws.merge_cells('D1:F1'); ws['D1'] = "Test Case Description" # Adjusted merge for data cell C1
    ws.merge_cells('A2:B2'); ws['A2'] = "Created By"
    ws.merge_cells('D2:F2'); ws['D2'] = "Reviewed By"           # Adjusted merge for data cell C2
    ws['G1'] = "Version"
    ws.merge_cells('B3:I3'); ws['A3'] = "QA Tester's Log"
    
    ws['C1'] = test_case_data['id']
    ws['G1'].value = test_case_data['description']
    ws.merge_cells('G1:I1')
    ws['C2'] = test_case_data['created_by']
    ws['G2'].value = test_case_data['reviewed_by']
    ws.merge_cells('G2:I2')
    ws['J1'].value = test_case_data['version']
    ws.merge_cells('J1:K1')
    ws['B3'] = test_case_data['qa_log']
    
    # Apply styles to header labels
    apply_header_style(ws['A1']); apply_header_style(ws['D1']); apply_header_style(ws['G1'])
    apply_header_style(ws['A2']); apply_header_style(ws['D2'])
    apply_header_style(ws['A3'])


    # Section 2: Tester Info
    ws.merge_cells('A5:B5'); ws['A5'] = "Tester's Name"
    ws.merge_cells('C5:D5'); ws['C5'] = test_case_data['tester_name']
    ws['E5'] = "Date Tested"
    ws.merge_cells('F5:G5'); ws['F5'] = test_case_data['date_tested']
    ws['H5'] = "Test Case (Pass/Fail/Not Executed)"
    ws.merge_cells('H5:I5')
    ws.merge_cells('J5:K5'); ws['J5'] = test_case_data['final_status']
    
    apply_header_style(ws['A5']); apply_header_style(ws['E5']); apply_header_style(ws['H5'])


    # Section 3: Prerequisites and Test Data
    ws['A7'] = "S #"; ws.merge_cells('B7:E7'); ws['B7'] = "Prerequisites"
    ws['F7'] = "S #"; ws.merge_cells('G7:K7'); ws['G7'] = "Test Data" # Merged to match image
    apply_header_style(ws['A7']); apply_header_style(ws['B7'])
    apply_header_style(ws['F7']); apply_header_style(ws['G7'])
    
    for i in range(4):
        ws[f'A{8+i}'] = i + 1
        ws.merge_cells(f'B{8+i}:E{8+i}')
        if i < len(test_case_data['prerequisites']):
            ws[f'B{8+i}'] = test_case_data['prerequisites'][i]
        
        ws[f'F{8+i}'] = i + 1
        ws.merge_cells(f'G{8+i}:K{8+i}')
        if i < len(test_case_data['test_data']):
            ws[f'G{8+i}'] = test_case_data['test_data'][i]
            
    # Section 4: Test Scenario
    ws.merge_cells('A13:B13'); ws['A13'] = "Test Scenario"
    ws.merge_cells('C13:K13'); ws['C13'] = test_case_data['scenario']
    apply_header_style(ws['A13'])

    # Section 5: Test Steps Table
    step_headers = ["Step #", "Step Details", "Expected Results", "Actual Results", "Pass / Fail / Not executed / Suspended"]
    ws.merge_cells('B15:D15'); ws.merge_cells('E15:F15'); ws.merge_cells('G15:I15'); ws.merge_cells('J15:K15')
    ws['A15'], ws['B15'], ws['E15'], ws['G15'], ws['J15'] = step_headers
    for cell_coord in ['A15', 'B15', 'E15', 'G15', 'J15']:
        apply_header_style(ws[cell_coord])
    
    start_row = 16
    for i, step in enumerate(test_case_data['steps']):
        row = start_row + i
        ws[f'A{row}'] = step['num']
        ws.merge_cells(f'B{row}:D{row}'); ws[f'B{row}'] = step['details']
        ws.merge_cells(f'E{row}:F{row}'); ws[f'E{row}'] = step['expected']
        ws.merge_cells(f'G{row}:I{row}'); ws[f'G{row}'] = step['actual']
        ws.merge_cells(f'J{row}:K{row}'); ws[f'J{row}'] = step['status']

    # Fill empty step rows for template look
    for i in range(len(test_case_data['steps']), 4):
        row = start_row + i
        ws[f'A{row}'] = i + 1
        ws.merge_cells(f'B{row}:D{row}')
        ws.merge_cells(f'E{row}:F{row}')
        ws.merge_cells(f'G{row}:I{row}')
        ws.merge_cells(f'J{row}:K{row}')

    # 5. --- Final Formatting ---
    # Re-apply data to top-left cells after merging
    ws['C1'] = test_case_data['id']
    ws['D1'].value = test_case_data['description']
    ws['C2'] = test_case_data['created_by']
    ws['D2'].value = test_case_data['reviewed_by']

    # Column Widths
    column_widths = {'A': 5, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 1, 'I': 15, 'J': 1, 'K': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Apply borders and alignment to all used cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=11):
        for cell in row:
            if cell.coordinate in ws.merged_cells or ws.cell(row=cell.row, column=cell.column).value is not None:
                 cell.border = thin_border
                 cell.alignment = alignment
                 
    # 6. --- Save the file ---
    filename = f"TestCase_{test_case_data['id']}.xlsx"
    wb.save(filename)
    print(f"✅ Successfully created '{filename}'")


if __name__ == "__main__":
    create_detailed_test_case_sheet()