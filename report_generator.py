#!/usr/bin/env python3
"""
Standalone Python script to convert a pytest JUnit XML report and an Excel-based
manual test report into a clean, professional, and human-readable Markdown
test case report with a main summary and detailed annexes.

This script can also export the final report to PDF using Pandoc.
"""

import xml.etree.ElementTree as ET
import argparse
from dataclasses import dataclass, field
from typing import List, Optional, Dict
from datetime import datetime
from collections import defaultdict
import sys
import re
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import subprocess
import shutil


# ==============================================================================
# 1. DATA MODELS (Streamlined for automation and manual tests)
# ==============================================================================

@dataclass
class TestStep:
    """A single step within a test case."""
    step_num: int
    details: str
    expected: str
    actual: str = "As Expected"
    status: str = "Pass"

@dataclass
class TestCase:
    """Represents a full, detailed test case, combining test results and metadata."""
    name: str
    time: float
    status: str
    failure_details: Optional[str] = None
    test_id: str = "N/A"
    description: str = "N/A"
    version: str = "1.0"
    tester_name: str = "Automation"
    date_tested: str = ""
    prerequisites: List[str] = field(default_factory=list)
    test_data: List[str] = field(default_factory=list)
    test_scenario: str = ""
    test_steps: List[TestStep] = field(default_factory=list)
    category: str = "Uncategorized"
    has_metadata: bool = False

# ==============================================================================
# 2. METADATA REGISTRY & CATEGORY MAPPING
# ==============================================================================
# This map defines the report sections based on test_id prefixes.
# Format: { "prefix": ("Display Name", sort_order) }
CATEGORY_MAP = {
    "MODEL_": ("Model & Unit Tests", 1),
    "API_":   ("API Tests", 2),
    "E2E_":   ("End-to-End (E2E) Tests", 3),
    "SCEN_":  ("Scenario & Integration Tests", 4),
}

# The single source of truth for all automated test case documentation.
METADATA_REGISTRY: Dict[str, Dict] = {
    # ==========================================================================
    # --- MODEL TESTS: app_user.py ---
    # Classname: tests.test-python.models.test_app_user
    # ==========================================================================
    "tests.test-python.models.test_app_user.test_create_user": {
        "test_id": "MODEL_USER_001",
        "description": "Verify AppUser model can create a user with a hashed password.",
        "test_scenario": "The create_user class method should correctly instantiate a user, hash their password, and add them to the database session.",
        "test_steps": [
            TestStep(1, "Call AppUser.create_user with valid details.", "An AppUser instance is returned."),
            TestStep(2, "Commit the session.", "The user is saved to the database with a user_id."),
            TestStep(3, "Verify the user's password.", "The verify_password method returns True for the correct password and False for an incorrect one."),
        ],
    },
    "tests.test-python.models.test_app_user.test_update_user": {
        "test_id": "MODEL_USER_002",
        "description": "Verify AppUser model can update user attributes.",
        "test_steps": [TestStep(1, "Create a user, then call update_user() with new data (e.g., last_name, balance).", "The user's attributes are updated in the database and the changes are persisted.")],
    },
    "tests.test-python.models.test_app_user.test_unique_email_constraint": {
        "test_id": "MODEL_DB_002",
        "description": "Verify database uniqueness constraint for user emails.",
        "test_scenario": "The database should prevent the creation of two users with the same email address, raising an IntegrityError.",
        "test_steps": [
            TestStep(1, "Create and commit User A with a specific email.", "User A is created successfully."),
            TestStep(2, "Create User B with the same email.", "User B instance is created in the session."),
            TestStep(3, "Attempt to commit the session with User B.", "The commit must fail and raise an IntegrityError."),
        ],
    },
    "tests.test-python.models.test_app_user.test_delete_user": {
        "test_id": "MODEL_USER_003",
        "description": "Verify AppUser model can delete a user.",
        "test_steps": [TestStep(1, "Create a user, retrieve their ID, then call delete_user().", "The user is removed from the database and can no longer be retrieved by their ID.")],
    },
    "tests.test-python.models.test_app_user.test_app_user_update_nothing_returns_false": {
        "test_id": "MODEL_USER_004",
        "description": "Verify calling update_user with no arguments returns False.",
        "test_steps": [TestStep(1, "On a user instance, call update_user() with no parameters.", "The method must return False, indicating no update was performed.")],
    },
    "tests.test-python.models.test_app_user.test_app_user_update_password_and_role": {
        "test_id": "MODEL_USER_005",
        "description": "Verify that updating password and role works correctly.",
        "test_steps": [
            TestStep(1, "Call update_user() with a new password and role.", "Method returns True, role is updated, and the new password can be verified while the old one cannot."),
        ],
    },
    "tests.test-python.models.test_app_user.test_app_user_update_to_existing_email_fails": {
        "test_id": "MODEL_DB_003",
        "description": "Verify that updating a user's email to an already existing email fails.",
        "test_steps": [TestStep(1, "Create two users. Attempt to update the second user's email to match the first user's email.", "The update_user() method must return False due to the unique constraint violation.")],
    },
    "tests.test-python.models.test_app_user.test_get_all_dicts": {
        "test_id": "MODEL_USER_006",
        "description": "Verify the get_all_dicts utility method for AppUser.",
        "test_steps": [TestStep(1, "Create multiple users and call AppUser.get_all_dicts().", "A list of dictionaries is returned, with each dictionary representing a user and containing the correct data (excluding password).")],
    },

    # --- MODEL TESTS: cafeteria.py ---
    "tests.test-python.models.test_cafeteria.test_create_cafeteria": { "test_id": "MODEL_CAFE_001", "description": "Verify Cafeteria model can create a new cafeteria.", "test_steps": [TestStep(1, "Call Cafeteria.create_cafeteria().", "A new cafeteria is created and persisted to the database.")] },
    "tests.test-python.models.test_cafeteria.test_update_cafeteria": { "test_id": "MODEL_CAFE_002", "description": "Verify Cafeteria model can update a cafeteria's name.", "test_steps": [TestStep(1, "Call update_cafeteria() with a new name.", "The cafeteria's name is successfully updated in the database.")] },
    "tests.test-python.models.test_cafeteria.test_get_all_dicts": { "test_id": "MODEL_CAFE_003", "description": "Verify get_all_dicts for Cafeteria returns all records.", "test_steps": [TestStep(1, "Create multiple cafeterias and call get_all_dicts().", "A list containing all created cafeteria dictionaries is returned.")] },
    "tests.test-python.models.test_cafeteria.test_delete_cafeteria": { "test_id": "MODEL_CAFE_004", "description": "Verify Cafeteria model can delete a cafeteria.", "test_steps": [TestStep(1, "Call delete_cafeteria() on an instance.", "The cafeteria is removed from the database.")] },
    "tests.test-python.models.test_cafeteria.test_update_cafeteria_with_no_data": { "test_id": "MODEL_CAFE_005", "description": "Verify update_cafeteria() with no arguments returns False.", "test_steps": [TestStep(1, "Call update_cafeteria() with no parameters.", "The method must return False.")] },
    
    # --- MODEL TESTS: daily_menu.py ---
    "tests.test-python.models.test_daily_menu.test_create_menu": { "test_id": "MODEL_MENU_001", "description": "Verify DailyMenu model can create a menu.", "test_steps": [TestStep(1, "Call DailyMenu.create_menu().", "A new menu is created and persisted.")] },
    "tests.test-python.models.test_daily_menu.test_update_menu": { "test_id": "MODEL_MENU_002", "description": "Verify DailyMenu model can update a menu.", "test_steps": [TestStep(1, "Call update_menu() with a new date.", "The menu's date is updated.")] },
    "tests.test-python.models.test_daily_menu.test_delete_menu": { "test_id": "MODEL_MENU_003", "description": "Verify DailyMenu model can delete a menu.", "test_steps": [TestStep(1, "Call delete_menu() on a menu instance.", "The menu is removed from the database.")] },
    "tests.test-python.models.test_daily_menu.test_update_menu_with_no_data": { "test_id": "MODEL_MENU_004", "description": "Verify update_menu() with no arguments returns False.", "test_steps": [TestStep(1, "Call update_menu() with no parameters.", "The method must return False.")] },
    "tests.test-python.models.test_daily_menu.test_menu_uniqueness_constraint_on_update": {
        "test_id": "MODEL_DB_001",
        "description": "Verify database uniqueness constraint for (cafeteria_id, menu_date) on update.",
        "test_scenario": "The system should prevent a menu from being updated to a date/cafeteria combination that already exists for another menu.",
        "test_steps": [
            TestStep(1, "Create Menu A for Cafeteria 1 on Date X.", "Menu is created."),
            TestStep(2, "Create Menu B for Cafeteria 2 on Date Y.", "Menu is created."),
            TestStep(3, "Attempt to update Menu B to use Cafeteria 1 and Date X.", "The update operation must fail and return False due to the unique constraint violation."),
        ]
    },
    "tests.test-python.models.test_daily_menu.test_get_all_menus_as_dicts": { "test_id": "MODEL_MENU_005", "description": "Verify get_all_dicts for DailyMenu returns all records.", "test_steps": [TestStep(1, "Create multiple menus and call get_all_dicts().", "A list containing all created menu dictionaries is returned.")] },
    
    # --- MODEL TESTS: daily_menu_item.py ---
    "tests.test-python.models.test_daily_menu_item.test_create_menu_item": { "test_id": "MODEL_MENUITEM_001", "description": "Verify a DailyMenuItem can be created.", "test_steps": [TestStep(1, "Call DailyMenuItem.create_menu_item().", "A new menu item is created.")] },
    "tests.test-python.models.test_daily_menu_item.test_update_menu_item": { "test_id": "MODEL_MENUITEM_002", "description": "Verify a DailyMenuItem can be updated.", "test_steps": [TestStep(1, "Call update_menu_item() with new data.", "The item's attributes are updated.")] },
    "tests.test-python.models.test_daily_menu_item.test_delete_menu_item": { "test_id": "MODEL_MENUITEM_003", "description": "Verify a DailyMenuItem can be deleted.", "test_steps": [TestStep(1, "Call delete_menu_item() on an instance.", "The item is removed.")] },
    "tests.test-python.models.test_daily_menu_item.test_update_menu_item_with_no_data": { "test_id": "MODEL_MENUITEM_004", "description": "Verify update_menu_item() with no arguments returns False.", "test_steps": [TestStep(1, "Call update_menu_item() with no parameters.", "The method must return False.")] },
    "tests.test-python.models.test_daily_menu_item.test_get_all_menu_items_as_dicts": { "test_id": "MODEL_MENUITEM_005", "description": "Verify get_all_dicts for DailyMenuItem returns all records.", "test_steps": [TestStep(1, "Create multiple items and call get_all_dicts().", "A list containing all created item dictionaries is returned.")] },

    # --- MODEL TESTS: dish.py ---
    "tests.test-python.models.test_dish.test_create_dish": { "test_id": "MODEL_DISH_001", "description": "Verify a Dish can be created.", "test_steps": [TestStep(1, "Call Dish.create_dish().", "A new dish is created.")] },
    "tests.test-python.models.test_dish.test_update_from_dict": { "test_id": "MODEL_DISH_002", "description": "Verify a Dish can be updated from a dictionary.", "test_steps": [TestStep(1, "Call update_from_dict() with new data.", "The dish's attributes are updated.")] },
    "tests.test-python.models.test_dish.test_get_by_id": { "test_id": "MODEL_DISH_003", "description": "Verify a Dish can be retrieved by its ID.", "test_steps": [TestStep(1, "Create a dish and call get_by_id() with its ID.", "The correct dish instance is returned.")] },
    "tests.test-python.models.test_dish.test_delete_dish": { "test_id": "MODEL_DISH_004", "description": "Verify a Dish can be deleted.", "test_steps": [TestStep(1, "Call delete_dish() on an instance.", "The dish is removed.")] },
    "tests.test-python.models.test_dish.test_get_all_dishes_as_dicts": { "test_id": "MODEL_DISH_005", "description": "Verify get_all_dicts for Dish returns all records.", "test_steps": [TestStep(1, "Create multiple dishes and call get_all_dicts().", "A list of all created dish dictionaries is returned.")] },
    "tests.test-python.models.test_dish.test_dish_delete_fails_if_in_use_by_menu_item": { "test_id": "MODEL_DB_004", "description": "Verify a dish cannot be deleted if referenced by a DailyMenuItem.", "test_steps": [TestStep(1, "Create a dish and add it to a menu.", "The setup is successful."), TestStep(2, "Attempt to delete the dish.", "The delete_dish() method must return False.")] },
    "tests.test-python.models.test_dish.test_dish_delete_fails_if_in_use_by_order_item": { "test_id": "MODEL_DB_005", "description": "Verify a dish cannot be deleted if referenced by an OrderItem.", "test_steps": [TestStep(1, "Create a dish and include it in a reservation.", "The setup is successful."), TestStep(2, "Attempt to delete the dish.", "The delete_dish() method must return False.")] },

    # --- MODEL TESTS: order_item.py ---
    "tests.test-python.models.test_order_item.test_create_order_item": { "test_id": "MODEL_ORDERITEM_001", "description": "Verify an OrderItem can be created.", "test_steps": [TestStep(1, "Call OrderItem.create_order_item().", "A new order item is created.")] },
    "tests.test-python.models.test_order_item.test_update_order_item": { "test_id": "MODEL_ORDERITEM_002", "description": "Verify an OrderItem can be updated.", "test_steps": [TestStep(1, "Call update_order_item() with new data.", "The item's attributes are updated.")] },
    "tests.test-python.models.test_order_item.test_delete_order_item": { "test_id": "MODEL_ORDERITEM_003", "description": "Verify an OrderItem can be deleted.", "test_steps": [TestStep(1, "Call delete_order_item() on an instance.", "The item is removed.")] },
    "tests.test-python.models.test_order_item.test_update_order_item_with_no_data": { "test_id": "MODEL_ORDERITEM_004", "description": "Verify update_order_item() with no arguments returns False.", "test_steps": [TestStep(1, "Call update_order_item() with no parameters.", "The method must return False.")] },
    "tests.test-python.models.test_order_item.test_get_all_order_items_as_dicts": { "test_id": "MODEL_ORDERITEM_005", "description": "Verify get_all_dicts for OrderItem returns all records.", "test_steps": [TestStep(1, "Create items and call get_all_dicts().", "A list of all created item dictionaries is returned.")] },

    # --- MODEL TESTS: reservation.py ---
    "tests.test-python.models.test_reservation.test_create_reservation": { "test_id": "MODEL_RESERVATION_001", "description": "Verify a Reservation can be created.", "test_steps": [TestStep(1, "Call Reservation.create_reservation().", "A new reservation is created.")] },
    "tests.test-python.models.test_reservation.test_update_reservation": { "test_id": "MODEL_RESERVATION_002", "description": "Verify a Reservation can be updated.", "test_steps": [TestStep(1, "Call update_reservation() with new data.", "The reservation's attributes are updated.")] },
    "tests.test-python.models.test_reservation.test_delete_reservation": { "test_id": "MODEL_RESERVATION_003", "description": "Verify a Reservation can be deleted.", "test_steps": [TestStep(1, "Call delete_reservation() on an instance.", "The reservation is removed.")] },
    "tests.test-python.models.test_reservation.test_update_reservation_with_no_data": { "test_id": "MODEL_RESERVATION_004", "description": "Verify update_reservation() with no arguments returns False.", "test_steps": [TestStep(1, "Call update_reservation() with no parameters.", "The method must return False.")] },
    "tests.test-python.models.test_reservation.test_get_all_reservations_as_dicts": { "test_id": "MODEL_RESERVATION_005", "description": "Verify get_all_dicts for Reservation returns all records.", "test_steps": [TestStep(1, "Create reservations and call get_all_dicts().", "A list of all created reservation dictionaries is returned.")] },

    # ==========================================================================
    # --- API TESTS ---
    # ==========================================================================
    "tests.test-python.controller.test_api_admin_auth.test_user_crud": {
        "test_id": "API_USER_001",
        "description": "Verify full CRUD (Create, Read, Update, Delete) for the User API endpoint.",
        "test_scenario": "An authenticated admin must be able to fully manage users through the API.",
        "prerequisites": ["An authenticated admin client."],
        "test_steps": [
            TestStep(1, "Send POST to `/api/v1/user/` to create a new user.", "Receives HTTP 201 and user data."),
            TestStep(2, "Send GET to `/api/v1/user/` to list all users.", "Receives HTTP 200 and the new user is in the list."),
            TestStep(3, "Send PUT to `/api/v1/user/{id}` to update the user.", "Receives HTTP 200 and updated data."),
            TestStep(4, "Send DELETE to `/api/v1/user/{id}` to delete the user.", "Receives HTTP 200."),
        ],
    },
    "tests.test-python.controller.test_api_admin_auth.test_cafeteria_crud": {
        "test_id": "API_CAFE_001",
        "description": "Verify full CRUD for the Cafeteria API endpoint.",
        "prerequisites": ["An authenticated admin client."],
        "test_steps": [TestStep(1, "Perform Create, Read, Update, Delete operations on the Cafeteria API.", "All operations succeed with correct HTTP status codes.")]
    },
    "tests.test-python.controller.test_api_admin_auth.test_dish_crud": {
        "test_id": "API_DISH_001",
        "description": "Verify full CRUD for the Dish API endpoint.",
        "prerequisites": ["An authenticated admin client."],
        "test_steps": [TestStep(1, "Perform Create, Read, Update, Delete operations on the Dish API.", "All operations succeed with correct HTTP status codes.")]
    },
    "tests.test-python.controller.test_api_admin_auth.test_menu_and_item_crud": {
        "test_id": "API_MENU_001",
        "description": "Verify creation of Menus and linking Menu Items via API.",
        "prerequisites": ["Authenticated admin client.", "A Cafeteria and a Dish exist in the DB."],
        "test_steps": [
            TestStep(1, "Create a DailyMenu via POST request.", "Receives HTTP 201 and menu data."),
            TestStep(2, "Create a DailyMenuItem via POST, linking the menu and a dish.", "Receives HTTP 201 and menu item data."),
            TestStep(3, "Delete the parent DailyMenu.", "The operation succeeds, and the child DailyMenuItem should be deleted by cascade."),
        ],
    },
    "tests.test-python.controller.test_api_auth.test_user_api_permissions": {
        "test_id": "API_SEC_001",
        "description": "Verify API permissions for a standard authenticated (non-admin) user.",
        "test_scenario": "A standard user should be able to access their own data but be denied access to other users' data or admin-only endpoints. This is a parametrized test that covers multiple endpoints.",
        "test_steps": [
            TestStep(1, "Attempt to access admin-only endpoints (e.g., list all users, create a dish).", "Access is denied with HTTP 401/403."),
            TestStep(2, "Attempt to access own user data (e.g., GET /api/v1/user/{own_id}).", "Access is allowed with HTTP 200."),
            TestStep(3, "Attempt to access another user's data (e.g., GET /api/v1/reservations/{other_user_reservation_id}).", "Access is denied with HTTP 401/403/404."),
        ]
    },
    "tests.test-python.controller.test_api_no_auth.test_api_unauthenticated_access_is_denied": {
        "test_id": "API_SEC_002",
        "description": "Verify that unauthenticated API access is denied for all protected endpoints.",
        "test_scenario": "An API client without a valid session cookie must be rejected with an HTTP 401 or 403 error when attempting to access any protected resource. This is a parametrized test that covers all protected API endpoints.",
        "prerequisites": ["Application is running."],
        "test_steps": [
            TestStep(1, "Send a request (GET, POST, PUT, DELETE) to a protected API endpoint without authentication.", "The server must respond with an HTTP 401 or 403 status code."),
        ],
    },

    # ==========================================================================
    # --- E2E TESTS ---
    # ==========================================================================
    "tests.test-python.selenium-e2e.test_loginAdminLogOut.TestLoginAdminLogOut.test_loginAdminLogOut": {
        "test_id": "E2E_LOGIN_001",
        "description": "Verify Admin Login and Logout functionality through the UI.",
        "prerequisites": ["Application is running.", "Default admin user exists."],
        "test_steps": [
            TestStep(1, "Navigate to login page.", "Page loads successfully."),
            TestStep(2, "Enter admin credentials and submit.", "Redirected to the admin dashboard."),
            TestStep(3, "Click the 'Logout' link.", "Redirected back to the login page."),
        ],
    },
    "tests.test-python.selenium-e2e.test_orderfoodinthepast.TestOrderfoodinthepast.test_orderfoodinthepast": {
        "test_id": "E2E_ORDER_001",
        "description": "Verify the end-to-end student workflow for ordering a meal.",
        "prerequisites": ["Application is running.", "Default student user exists.", "Menus are available for the current month."],
        "test_scenario": "A student logs in, navigates to a cafeteria, selects a date, adds a dish to their cart, places the order, and verifies the order in their history.",
        "test_steps": [
            TestStep(1, "Navigate to login page and log in as a student.", "Redirected to user dashboard."),
            TestStep(2, "Navigate between different cafeteria links.", "The main content updates to show the selected cafeteria's menu page."),
            TestStep(3, "Open the date picker and select a date.", "The menu table updates to show dishes for the selected date."),
            TestStep(4, "Add an available item to the cart.", "The item appears in the cart summary and the 'Add' button changes to 'Added'."),
            TestStep(5, "Click 'Place Order'.", "A success notification appears and the user is redirected to the 'Order History' page."),
            TestStep(6, "Verify the new order appears on the 'Order History' page.", "The order history list contains the newly placed order."),
        ],
    },
    
    # ==========================================================================
    # --- SCENARIO TESTS: test_combined_scenarios.py ---
    # ==========================================================================
    "tests.test-python.selenium-e2e.test_combined_scenarios.TestUserManagementScenario.test_complete_user_lifecycle": {
        "test_id": "SCEN_USER_001",
        "description": "Test the full lifecycle of user management from creation to deletion at the model level.",
        "test_steps": [TestStep(1, "Execute create, update, and delete functions on the AppUser model, including constraint checks.", "All model methods execute without error and reflect correct state changes in the database.")]
    },
    "tests.test-python.selenium-e2e.test_combined_scenarios.TestCafeteriaManagementScenario.test_complete_cafeteria_lifecycle": {
        "test_id": "SCEN_CAFE_001",
        "description": "Test the full lifecycle of cafeteria management at the model level.",
        "test_steps": [TestStep(1, "Execute create, update, and delete functions on the Cafeteria model.", "All model methods execute without error and reflect correct state changes in the database.")]
    },
    "tests.test-python.selenium-e2e.test_combined_scenarios.TestMenuManagementScenario.test_complete_menu_creation_workflow": {
        "test_id": "SCEN_MENU_001",
        "description": "Test the full menu and dish lifecycle, including referential integrity.",
        "test_scenario": "Create all related entities (Cafeteria, Dish, Menu, MenuItem), then verify that an in-use dish cannot be deleted. Finally, clean up all entities in the correct order.",
        "test_steps": [
            TestStep(1, "Create a Cafeteria, Dish, DailyMenu, and DailyMenuItem.", "All entities are created successfully."),
            TestStep(2, "Attempt to delete the Dish while it is linked to the DailyMenuItem.", "The deletion must fail, returning False."),
            TestStep(3, "Delete the DailyMenuItem first, then delete the Dish.", "Both deletions must now succeed."),
        ]
    },
    "tests.test-python.selenium-e2e.test_combined_scenarios.TestOrderWorkflowScenario.test_complete_order_workflow": {
        "test_id": "SCEN_ORDER_001",
        "description": "Test the complete order processing workflow at the model level.",
        "test_steps": [TestStep(1, "Execute model functions for creating users, dishes, menus, reservations, and order items in sequence.", "All model methods execute without error, simulating a full order process.")]
    },
    "tests.test-python.selenium-e2e.test_combined_scenarios.TestSecurityAndAuthScenario.test_unauthenticated_access_security": {
        "test_id": "SCEN_SEC_001",
        "description": "Verify that a sample of protected API endpoints reject unauthenticated access.",
        "test_steps": [TestStep(1, "Call a subset of protected API endpoints without a login session.", "All calls must be rejected with a 401 or 403 status code.")]
    },
    "tests.test-python.selenium-e2e.test_combined_scenarios.TestDataIntegrityScenario.test_referential_integrity_workflow": {
        "test_id": "SCEN_DB_001",
        "description": "Verify multiple data integrity constraints across the database schema.",
        "test_steps": [
            TestStep(1, "Test AppUser email uniqueness constraint.", "The database prevents duplicate emails."),
            TestStep(2, "Test that a Dish in use by a menu cannot be deleted.", "The deletion operation fails as expected."),
            TestStep(3, "Perform cleanup in the correct order to respect foreign key constraints.", "All entities are deleted successfully without integrity errors."),
        ]
    },
    "tests.test-python.selenium-e2e.test_combined_scenarios.TestSystemIntegrationScenario.test_full_system_integration": {
        "test_id": "SCEN_INT_001",
        "description": "Test the full system integration by combining model and API tests in a realistic sequence.",
        "test_steps": [
            TestStep(1, "Phase 1: System Setup (create users, cafeterias, dishes).", "All base data is created via model methods."),
            TestStep(2, "Phase 2: Menu Management (create menus and items).", "Menus are created successfully."),
            TestStep(3, "Phase 3 & 4: Order Processing and Management.", "Orders are created and updated successfully."),
            TestStep(4, "Phase 5: Data Management (update existing data).", "Updates are successful."),
            TestStep(5, "Phase 6: Security Verification (check unauthenticated access).", "Access to a protected endpoint is denied."),
            TestStep(6, "Phase 7: System Cleanup.", "All created data is successfully deleted in the correct order."),
        ]
    },
}

# ==============================================================================
# 3. UTILITY FUNCTIONS
# ==============================================================================

def create_anchor_id(test_id: str) -> str:
    """Creates a URL-friendly anchor ID from a test case ID."""
    return f"test-case-{re.sub(r'[^a-z0-9]+', '-', test_id.lower())}"

def export_to_pdf(md_input_path: str, pdf_output_path: str) -> bool:
    """
    Exports a Markdown file to PDF using Pandoc.

    This function requires Pandoc and a LaTeX distribution (like MiKTeX,
    TeX Live, or MacTeX) to be installed and available in the system's PATH.
    """
    if not shutil.which("pandoc"):
        print("Error: 'pandoc' command not found.", file=sys.stderr)
        print("Please install Pandoc from https://pandoc.org/installing.html to use PDF export.", file=sys.stderr)
        return False

    if not shutil.which("pdflatex"):
        print("Warning: 'pdflatex' command not found. PDF generation may fail.", file=sys.stderr)
        print("Please install a LaTeX distribution (e.g., MiKTeX, TeX Live) for robust PDF support.", file=sys.stderr)

    pandoc_command = [
        "pandoc",
        md_input_path,
        "-o",
        pdf_output_path,
        "--from=markdown",
        "--pdf-engine=pdflatex",
        "--toc",  # Add a table of contents
        "--number-sections",  # Number the sections
        "--number-offset=-1", # Start numbering at level 2 headers (##)
        "-V", "geometry:a4paper,margin=1in",  # Set paper size and margins
        "-V", "colorlinks",  # Make links colored instead of boxed
    ]

    try:
        print(f"Running command: {' '.join(pandoc_command)}")
        subprocess.run(
            pandoc_command,
            check=True,
            capture_output=True,
            text=True,
            encoding='utf-8'
        )
        return True
    except FileNotFoundError:
        # This is a fallback, shutil.which should have caught it.
        print("Error: 'pandoc' command not found. Could not execute the process.", file=sys.stderr)
        return False
    except subprocess.CalledProcessError as e:
        print(f"Error: Pandoc failed to convert the file to PDF.", file=sys.stderr)
        print(f"Pandoc Exit Code: {e.returncode}", file=sys.stderr)
        print("\n--- Pandoc STDOUT ---", file=sys.stderr)
        print(e.stdout, file=sys.stderr)
        print("\n--- Pandoc STDERR ---", file=sys.stderr)
        print(e.stderr, file=sys.stderr)
        print("\nEnsure a full LaTeX distribution is installed and configured correctly.", file=sys.stderr)
        return False


# ==============================================================================
# 4. PARSING LOGIC
# ==============================================================================

def parse_automated_tests_from_xml(xml_file: str) -> List[TestCase]:
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        suite_element = root.find('testsuite')
    except (ET.ParseError, FileNotFoundError, AttributeError) as e:
        print(f"Error: Could not parse or find a valid <testsuite> in '{xml_file}'.\nDetails: {e}", file=sys.stderr)
        sys.exit(1)

    timestamp = suite_element.get('timestamp', datetime.now().isoformat())
    all_test_cases = []

    for case_element in suite_element.findall('testcase'):
        name = case_element.get('name', 'Unknown Test')
        classname = case_element.get('classname', 'unknown_test_suite')
        time = float(case_element.get('time', '0'))
        status, failure_details = 'Passed', None

        if case_element.find('failure') is not None:
            status = 'Failed'
            failure_details = case_element.find('failure').text
        elif case_element.find('error') is not None:
            status = 'Failed' # Treat errors as failures for reporting
            failure_details = case_element.find('error').text
        elif case_element.find('skipped') is not None:
            status = 'Skipped'

        case = TestCase(name=name, time=time, status=status, failure_details=failure_details)
        
        # Match parametrized tests like 'test_function[param]'
        clean_name = name.split('[')[0]
        registry_key = f"{classname}.{clean_name}"

        if registry_key in METADATA_REGISTRY:
            metadata = METADATA_REGISTRY[registry_key]
            case.has_metadata = True
            for key, value in metadata.items():
                if hasattr(case, key):
                    setattr(case, key, value)
            
            for prefix, (category_name, _) in CATEGORY_MAP.items():
                if case.test_id.startswith(prefix):
                    case.category = category_name
                    break

        case.date_tested = datetime.fromisoformat(timestamp).strftime('%d-%b-%Y')
        
        if case.status in ('Failed', 'Error') and case.test_steps:
            last_step = case.test_steps[-1]
            last_step.status = "Failed"
            last_step.actual = f"Execution failed. See details below."

        all_test_cases.append(case)

    return all_test_cases

def parse_manual_tests_from_excel(excel_file: str) -> List[TestCase]:
    """Parses manual test cases from the specified Excel file."""
    if not excel_file:
        return []
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except (FileNotFoundError, InvalidFileException) as e:
        print(f"Warning: Could not open or parse Excel file '{excel_file}'. Manual tests will be skipped.\nDetails: {e}", file=sys.stderr)
        return []

    if "Test Summary" not in wb.sheetnames:
        print("Warning: 'Test Summary' sheet not found in Excel file. Manual tests will be skipped.", file=sys.stderr)
        return []

    summary_ws = wb["Test Summary"]
    manual_cases = []
    
    header = [cell.value for cell in summary_ws[1]]
    try:
        # Find column indices dynamically
        id_col_idx = header.index("Test ID")
        title_col_idx = header.index("Title")
        status_col_idx = header.index("Status")
        tester_col_idx = header.index("Assigned To")
        date_col_idx = header.index("Date Tested")
    except ValueError as e:
        print(f"Warning: A required column ({e}) not found in 'Test Summary' sheet. Skipping manual tests.", file=sys.stderr)
        return []

    for row in summary_ws.iter_rows(min_row=2, values_only=True):
        test_id = row[id_col_idx]
        if not test_id or test_id not in wb.sheetnames:
            continue

        ws = wb[test_id]
        
        # Helper to safely get value from potentially merged cells
        def get_cell_value(cell_coord):
            cell = ws[cell_coord]
            # If it's a merged cell, find the top-left master cell and get its value
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            return cell.value
        
        status_raw = str(row[status_col_idx] or 'Not Executed').title()
        final_status = 'Passed' if status_raw == 'Pass' else status_raw
        
        case = TestCase(
            name=get_cell_value('G1') or row[title_col_idx],
            time=0.0,
            test_id=test_id,
            description=get_cell_value('C13') or "N/A", # Test Scenario
            tester_name=row[tester_col_idx],
            version=str(get_cell_value('J2') or "1.0"),
            status=final_status,
            failure_details=get_cell_value('B3'), # QA Tester's Log for notes
            has_metadata=True,
            category="Manual Tests"
        )
        
        dt_val = row[date_col_idx]
        if isinstance(dt_val, datetime):
            case.date_tested = dt_val.strftime('%d.%m.%Y')
        elif isinstance(dt_val, str):
            case.date_tested = dt_val
        elif dt_val:
            case.date_tested = str(dt_val)

        case.prerequisites = [get_cell_value(f'B{r}') for r in range(8, 12) if get_cell_value(f'B{r}')]

        case.test_steps = []
        for i in range(16, 26): # Check more rows for steps
            step_num = get_cell_value(f'A{i}')
            step_details = get_cell_value(f'B{i}')
            if step_num and step_details:
                case.test_steps.append(TestStep(
                    step_num=step_num,
                    details=step_details,
                    expected=get_cell_value(f'E{i}'),
                    actual=get_cell_value(f'G{i}'),
                    status=str(get_cell_value(f'J{i}') or "").title()
                ))
        
        manual_cases.append(case)

    return manual_cases


# ==============================================================================
# 5. MARKDOWN GENERATION LOGIC
# ==============================================================================

def generate_global_summary(automated_cases: List[TestCase], manual_cases: List[TestCase]) -> str:
    """Generates a high-level summary table of all test results."""
    
    stats = defaultdict(lambda: defaultdict(int))
    
    for case in automated_cases:
        stats[case.category]['total'] += 1
        stats[case.category][case.status.lower()] += 1
        
    for case in manual_cases:
        stats[case.category]['total'] += 1
        stats[case.category][case.status.lower()] += 1
        
    rows = []
    total_stats = defaultdict(int)
    
    sorted_categories = sorted(
        [cat for cat in stats.keys() if cat != "Manual Tests"],
        key=lambda cat_name: next((c[1] for c in CATEGORY_MAP.values() if c[0] == cat_name), 99)
    )
    if "Manual Tests" in stats:
        sorted_categories.append("Manual Tests")

    for category in sorted_categories:
        s = stats[category]
        total = s['total']
        passed = s.get('passed', 0)
        failed = s.get('failed', 0) + s.get('error', 0) + s.get('fail', 0)
        skipped = s.get('skipped', 0)
        
        total_stats['total'] += total
        total_stats['passed'] += passed
        total_stats['failed'] += failed
        total_stats['skipped'] += skipped
        
        pass_rate = (passed / (total - skipped) * 100) if (total - skipped) > 0 else 0
        
        row = f"| {category} | {total} | {passed} | {failed} | {skipped} | {pass_rate:.1f}% |"
        rows.append(row)

    total_pass_rate = (total_stats['passed'] / (total_stats['total'] - total_stats['skipped']) * 100) if (total_stats['total'] - total_stats['skipped']) > 0 else 0
    total_row = f"| **Total** | **{total_stats['total']}** | **{total_stats['passed']}** | **{total_stats['failed']}** | **{total_stats['skipped']}** | **{total_pass_rate:.1f}%** |"

    header = "| Test Category | Total | Passed | Failed | Skipped | Pass Rate |\n| :--- | :---: | :---: | :---: | :---: | :---: |"
    return f"{header}\n" + "\n".join(rows) + "\n|---|---|---|---|---|---|\n" + total_row

def generate_failed_tests_summary(all_cases: List[TestCase]) -> str:
    """Creates a summary of only the failed tests with links to the annex."""
    failed_cases = [case for case in all_cases if case.status.lower() in ('failed', 'error', 'fail')]
    
    if not failed_cases:
        return "All tests passed. No failures to report."

    rows = []
    for case in sorted(failed_cases, key=lambda c: c.test_id):
        anchor = create_anchor_id(case.test_id)
        row = f"| `{case.test_id}` | {case.name} | {case.category} | [See Details](#{anchor}) |"
        rows.append(row)

    # Give the 'Description' column much more relative width to encourage text wrapping
    header = "| Test ID | Description | Category | Details |\n| :--- | :--------------------------------- | :--- | :---: |"
    return f"{header}\n" + "\n".join(rows)

def generate_annex_automated_report(all_cases: List[TestCase]) -> str:
    """Generates the full, detailed report for automated tests for the annex."""
    if not all_cases:
        return "No automated test cases were found in the provided XML file."
        
    documented_cases = [case for case in all_cases if case.has_metadata]
    undocumented_cases = [case for case in all_cases if not case.has_metadata]
    categorized_cases = defaultdict(list)
    for case in documented_cases:
        categorized_cases[case.category].append(case)

    report_parts = ["## Annex A: Automated Test Execution Details\n"]
    
    total = len(all_cases)
    passed = sum(1 for tc in all_cases if tc.status == 'Passed')
    failed = sum(1 for tc in all_cases if tc.status in ('Failed', 'Error'))
    skipped = sum(1 for tc in all_cases if tc.status == 'Skipped')
    summary_table = f"""
| Metric | Value |
|---|---|
| **Total Automated Tests** | {total} |
| Passed | {passed} |
| Failed/Error | {failed} |
| Skipped | {skipped} |
"""
    report_parts.append(summary_table)

    sorted_category_names = sorted(
        [cat for cat in categorized_cases.keys()],
        key=lambda cat_name: next((c[1] for c in CATEGORY_MAP.values() if c[0] == cat_name), 99)
    )

    for category_name in sorted_category_names:
        report_parts.append(f"\n### {category_name}\n")
        cases_in_category = sorted(categorized_cases[category_name], key=lambda c: (c.status != 'Passed', c.test_id))
        for case in cases_in_category:
            report_parts.append(generate_detailed_automated_case_markdown(case))
            
    if undocumented_cases:
        report_parts.append("\n### Undocumented Test Cases\n")
        report_parts.append("The following test cases were executed but have no metadata in the registry. Consider documenting them.\n")
        # Give 'Test Function' column more width
        report_parts.append("\n| Test Function | Status | Time |\n|:---------------------------------|:---|:---|")
        for case in sorted(undocumented_cases, key=lambda c: c.name):
            report_parts.append(f"| `{case.name}` | {case.status} | {case.time:.4f}s |")
        
    return "\n".join(report_parts)

def generate_detailed_automated_case_markdown(case: TestCase) -> str:
    """Generates the Markdown for a single, detailed automated test case."""
    anchor_id = create_anchor_id(case.test_id)
    header = f"""
#### <a id="{anchor_id}"></a>{case.test_id}: {case.description}
| Attribute | Value |
| :--- | :--- |
| **Test Case ID** | `{case.test_id}` |
| **Version** | `{case.version}` |
| **Tester** | `{case.tester_name}` |
| **Execution Time** | `{case.time:.4f}s` |
| **Date Tested** | `{case.date_tested}` |
| **Final Status** | **{case.status}** |
| **Test Function**| `{case.name}` |
"""
    
    prereqs = "\n".join([f"1. {p}" for p in case.prerequisites]) or "None"
    
    details_section = f"""
**Prerequisites:**
{prereqs}

**Test Scenario:**
{case.test_scenario or "N/A"}
"""

    steps_content = "*No detailed steps defined in metadata.*"
    if case.test_steps:
        # Give relevant columns more width for wrapping
        steps_header = "\n| Step # | Step Details | Expected Results | Actual Results | Status |\n|:---:|:----------------------|:----------------------|:----------------------|:---:|\n"
        steps_rows = "\n".join(
            [f"| {s.step_num} | {s.details} | {s.expected} | {s.actual} | **{s.status}** |" for s in case.test_steps]
        )
        steps_content = steps_header + steps_rows

    failure_block = ""
    if case.failure_details:
        details = case.failure_details.strip()
        failure_block = f"\n**Failure Details:**\n```\n{details}\n```\n"

    return f"{header}\n{details_section}\n**Test Steps:**\n{steps_content}\n{failure_block}\n---"

def generate_annex_manual_report(manual_cases: List[TestCase]) -> str:
    """Generates the detailed report for manual tests for the annex."""
    if not manual_cases:
        return "## Annex B: Manual Test Execution Details\n\nNo manual test cases were found or executed."

    report_parts = ["## Annex B: Manual Test Execution Details\n"]
    sorted_cases = sorted(manual_cases, key=lambda c: c.test_id)
    
    for case in sorted_cases:
        report_parts.append(generate_single_manual_case_markdown(case))
        
    return "\n".join(report_parts)

def generate_single_manual_case_markdown(case: TestCase) -> str:
    """Generates the Markdown for a single manual test case from Excel data."""
    anchor_id = create_anchor_id(case.test_id)
    header = f"""
#### <a id="{anchor_id}"></a>Manual Test Case: {case.test_id}

| Attribute | Value |
| :--- | :--- |
| **Test Case ID** | `{case.test_id}` |
| **Title** | {case.name} |
| **Description** | {case.description} |
| **Tester** | {case.tester_name} |
| **Date Tested** | {case.date_tested} |
"""
    
    prereqs = "\n".join([f"1.  {p}" for p in case.prerequisites]) or "None"
    prereqs_section = f"""
**Prerequisites:**
{prereqs}
"""
    
    # Give relevant columns more width for wrapping
    steps_header = "\n| Step # | Action | Expected Result | Actual Result | Status |\n| :--- | :---------------------- | :---------------------- | :---------------------- | :--- |\n"
    
    # Ensure status is properly formatted or empty
    def format_status(s):
        if not s or s.lower() == 'none':
            return ''
        return s.title()
        
    steps_rows = "\n".join(
        [f"| {s.step_num} | {s.details} | {s.expected} | {s.actual or ''} | {format_status(s.status)} |" for s in case.test_steps]
    )
    steps_content = steps_header + steps_rows

    final_status = case.status
    if not final_status or final_status.lower() == 'none':
        final_status = 'NOT EXECUTED'
        
    final_result = f"**Final Result:** **{final_status.upper()}**\n"
    notes = f"**Notes/Comments:**\n{case.failure_details}\n" if case.failure_details else ""

    return f"{header}\n{prereqs_section}\n**Test Steps:**{steps_content}\n\n{final_result}\n{notes}\n---"


# ==============================================================================
# 6. MAIN EXECUTION
# ==============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert JUnit XML and an Excel manual test report to a clean Markdown report with a summary and annexes.",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument("input_xml", help="Path to the input JUnit XML file.")
    parser.add_argument("output_file", help="Path for the final output Markdown file.")
    parser.add_argument("--template", help="Optional path to a Markdown template file with placeholders.")
    parser.add_argument("--manual-tests-excel", help="Optional path to an Excel file with manual test cases.")
    parser.add_argument("--pdf-output", help="Optional path to export the final report as a PDF using Pandoc.\nRequires Pandoc and a LaTeX distribution to be installed.")
    
    args = parser.parse_args()
    
    print(f"Parsing automated tests from XML: {args.input_xml}...")
    automated_cases = parse_automated_tests_from_xml(args.input_xml)
    print(f"Found {len(automated_cases)} automated test cases.")

    manual_cases = []
    if args.manual_tests_excel:
        print(f"Parsing manual tests from Excel: {args.manual_tests_excel}...")
        manual_cases = parse_manual_tests_from_excel(args.manual_tests_excel)
        print(f"Found {len(manual_cases)} manual test cases.")
    else:
        print("No manual test Excel file provided, skipping.")

    print("Generating Markdown report content...")
    all_cases = automated_cases + manual_cases
    
    # Generate the main body content
    summary_md = generate_global_summary(automated_cases, manual_cases)
    failed_tests_md = generate_failed_tests_summary(all_cases)
    
    # Generate the annex content
    annex_automated_md = generate_annex_automated_report(automated_cases)
    annex_manual_md = generate_annex_manual_report(manual_cases)
    
    final_content = ""
    
    if args.template:
        print(f"Template mode enabled. Reading template: {args.template}")
        try:
            with open(args.template, "r", encoding="utf-8") as f:
                template_content = f.read()
        except FileNotFoundError:
            print(f"Error: Template file not found at '{args.template}'", file=sys.stderr)
            sys.exit(1)

        final_content = template_content
        # Replace placeholders for the main body
        final_content = final_content.replace("%%TEST_SUMMARY%%", summary_md)
        final_content = final_content.replace("%%FAILED_TESTS_SUMMARY%%", failed_tests_md)
        
        # Replace placeholders for the annexes
        annex_content = f"{annex_automated_md}\n\n{annex_manual_md}"
        final_content = final_content.replace("%%TEST_ANNEX%%", annex_content)
        
    else:
        # Standalone mode (if no template is provided)
        final_content = (
            f"# Test Execution Report\n\n"
            f"## Overall Test Statistics\n{summary_md}\n\n"
            f"## Key Findings (Failures)\n{failed_tests_md}\n\n"
            f"---\n\n"
            f"# Annexes\n\n"
            f"{annex_automated_md}\n\n"
            f"{annex_manual_md}"
        )

    try:
        with open(args.output_file, "w", encoding="utf-8") as f:
            f.write(final_content)
        print(f"✅ Successfully generated Markdown report at: {args.output_file}")
    except IOError as e:
        print(f"Error: Could not write to output file '{args.output_file}'.\n{e}", file=sys.stderr)
        sys.exit(1)

    # Optional: Export to PDF using Pandoc
    if args.pdf_output:
        print(f"\nAttempting to export Markdown to PDF: {args.pdf_output}...")
        if export_to_pdf(args.output_file, args.pdf_output):
            print(f"✅ Successfully exported PDF report to: {args.pdf_output}")
        else:
            print(f"❌ Failed to export PDF. Please check the errors above.", file=sys.stderr)

    if any(not case.has_metadata for case in automated_cases):
        undocumented_count = sum(1 for case in automated_cases if not case.has_metadata)
        print(f"\n⚠️  Warning: {undocumented_count} automated test cases were found without metadata.")